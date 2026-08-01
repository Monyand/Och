import openpyxl

file_path = r'd:\ОЧ\Програми\УСІ_ПЕРЕВІРКИ_01.08_логістика.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheets:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n--- SHEET: {sheetname} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
    for r in range(1, min(ws.max_row + 1, 30)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")
